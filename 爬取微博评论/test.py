from selenium import webdriver
import time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
driver = webdriver.Chrome("../chromedriver.exe")
# 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
wb = Workbook()
# 获取当前活跃的worksheet,默认就是第一个worksheet
ws = wb.active
ws.cell(row=1, column=1).value = "序号"
ws.cell(row=1, column=2).value = "评论"
# 登陆weibo
driver.maximize_window()
driver.implicitly_wait(5)  # 元素等待
driver.get("http://www.weibo.com/login.php")
driver.find_element_by_id("loginname").clear()
driver.find_element_by_id("loginname").send_keys("18998634187")
driver.find_element_by_name("password").send_keys("q1937838")
time.sleep(2)
# 带空格的没办法找到元素，所以先找到上面的没有空格的元素，然后找子元素
driver.find_element_by_xpath("//div[@class='W_login_form']//div[6]//a[1]").click()
# 模拟输入关键词
time.sleep(10)
# 滴滴 等车 司机 分钟
# 等车 车程 分钟
driver.find_elements_by_css_selector(".W_input")[0].send_keys("滴滴 拼车 分钟")
time.sleep(1)
# file = open("res.txt", 'w', encoding='utf-8')
driver.find_element_by_css_selector("[title='搜索']").click()
cnt = 0
for i in range(42):
    data = driver.find_elements_by_css_selector("[node-type='feed_list_content']")
    for item in data:
        # value_of_css_property('color')
        # text
        try:
            sentence = str(item.text).strip()
            # file.write(str(item.text).strip())
            # file.write("\n\n")
            # ws.cell(row=cnt+1, column=1).value = str(cnt+1)
            # ws.cell(row=cnt+1, column=2).value = sentence
            ws.append([str(cnt+1), sentence])
            cnt += 1
        except:
            print("编码解码错误")
        print(item.text)
    try:
        driver.find_elements_by_class_name('next')[0].click()
    except:
        break
    time.sleep(4)
# file.close()
wb.save("滴滴拼车2.xlsx")

# driver.find_element_by_name()
# 通过id定位
# driver.find_element_by_id('kw')
# 通过name定位
# driver.find_element_by_name('wd')
# 通过class定位
# driver.find_element_by_class_name("lala")
# 通过tag定位
# driver.find_element_by_tag_name('input')
# 通过css定位
# driver.find_element_by_css_selector('#kw')
# 通过xpath定位
# driver.find_element_by_xpath('//form[@id="form"]/span/input[@id="kw"]')
# 通过link定位、
# driver.find_element_by_link_text('贴吧')

